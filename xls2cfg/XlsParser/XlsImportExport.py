#coding: utf-8
#@author sundream
#@date 2026-08-07
"""Excel <-> schema/json/binary round-trip (MapEditor).

- write_data_from_xlsx: xlsx -> {output}/schema + {output}/json
- write_xlsx_from_data / write_xlsx_from_dir: schema + json|binary -> xlsx
- field.layout restores merge-split columns (P1 keys / P2 element-per-column)
"""

from __future__ import annotations

import json
import os
import re
import shutil
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    openpyxl = None
    load_workbook = None

from XlsParser.Config import Config
from XlsParser.Sheet import Sheet, is_importable_sheet_title, split_at_display
from XlsParser.Xls2JsonParser import Xls2JsonParser
from XlsParser.Xls2SchemaParser import (
    Xls2SchemaParser,
    KIND_CLASS,
    KIND_1D,
    KIND_2D,
)
from XlsParser.ByteStream import ByteStream
from XlsParser.Type import Type, parse_tags_cell, format_tags_cell
from XlsParser.XlsClass import readClass
from XlsParser.XlsEnum import readEnum

CLASS_WORKBOOK = "__class__.xlsx"
ENUM_WORKBOOK = "__enum__.xlsx"
CLASS_JSON = "__class__.json"
ENUM_JSON = "__enum__.json"
TYPE_BUNDLE_FILES = frozenset((CLASS_JSON, ENUM_JSON))


def _write_json_list(path, data):
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    text = json.dumps(data, ensure_ascii=False, indent=2)
    with open(path, "w", encoding="utf-8", newline="\n") as fd:
        fd.write(text)
        if not text.endswith("\n"):
            fd.write("\n")
    print("write", path)
    return path


def _copy_json(src, dest):
    parent = os.path.dirname(dest)
    if parent:
        os.makedirs(parent, exist_ok=True)
    if os.path.abspath(src) != os.path.abspath(dest):
        shutil.copyfile(src, dest)
    print("write", dest)
    return dest


def _class_defs_to_json_list(defs):
    from XlsParser.Type import Field as TypeField
    out = []
    for typename, entry in (defs or {}).items():
        item = {"typename": typename}
        if entry.get("displayName"):
            item["displayName"] = entry["displayName"]
        fields = []
        for f in entry.get("fields") or []:
            if not f.get("name") or not f.get("type"):
                continue
            fo = {"name": f["name"], "type": f["type"]}
            if f.get("comment"):
                fo["displayName"] = f["comment"]
            tags = TypeField.format_tags(f.get("tags"))
            if tags:
                fo["tags"] = tags
            if f.get("group"):
                fo["group"] = f["group"]
            if f.get("remarks"):
                fo["remarks"] = f["remarks"]
            fields.append(fo)
        item["fields"] = fields
        out.append(item)
    return out


def _enum_defs_to_json_list(defs):
    from XlsParser.Type import Field as TypeField
    out = []
    for typename, entry in (defs or {}).items():
        item = {
            "typename": typename,
            "enumType": entry.get("enumType") or "int32",
        }
        if entry.get("displayName"):
            item["displayName"] = entry["displayName"]
        if entry.get("flags"):
            item["flags"] = True
        else:
            item["flags"] = False
        item["exportType"] = bool(entry.get("exportType"))
        fields = []
        for f in entry.get("fields") or []:
            if not f.get("name"):
                continue
            fo = {"name": f["name"], "value": f.get("value")}
            if f.get("comment"):
                fo["displayName"] = f["comment"]
            tags = TypeField.format_tags(f.get("tags"))
            if tags:
                fo["tags"] = tags
            fields.append(fo)
        item["fields"] = fields
        out.append(item)
    return out


def _remove_kind0_schemas_for_workbook(schema_dir, workbook_name):
    if not os.path.isdir(schema_dir):
        return
    for fn in os.listdir(schema_dir):
        if not fn.endswith(".json") or fn in TYPE_BUNDLE_FILES:
            continue
        path = os.path.join(schema_dir, fn)
        try:
            obj = json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception:
            continue
        if not isinstance(obj, dict) or obj.get("kind") != KIND_CLASS:
            continue
        wb = os.path.basename(obj.get("workbook") or "")
        if wb == workbook_name:
            os.remove(path)
            print("remove", path)


def _merge_type_defs(collect_xlsx, collect_json, basename, excel_dir, schema_json, schema_wins):
    """Merge type defs from local xlsx/json and optional schema json.

    schema_wins=True: local xlsx, local json, then schema json (export).
    schema_wins=False: schema json, local xlsx, then local json (import_xlsx).
    """
    from XlsParser.TypeDefPath import resolve_type_def_paths

    xlsx, js = resolve_type_def_paths(excel_dir, basename) if excel_dir else (None, None)
    prev = Config.tags
    Config.tags = []
    try:
        xlsx_defs = collect_xlsx(xlsx, apply_tags_filter=False) if xlsx else OrderedDict()
        json_defs = collect_json(js, apply_tags_filter=False) if js else OrderedDict()
        schema_defs = (
            collect_json(schema_json, apply_tags_filter=False)
            if schema_json and os.path.isfile(schema_json) else OrderedDict()
        )
    finally:
        Config.tags = prev
    defs = OrderedDict()
    if schema_wins:
        defs.update(xlsx_defs)
        defs.update(json_defs)
        defs.update(schema_defs)
    else:
        defs.update(schema_defs)
        defs.update(xlsx_defs)
        defs.update(json_defs)
    return defs


def _write_type_bundle_json(json_list, json_path):
    results = []
    if json_path and json_list:
        _write_json_list(json_path, json_list)
        results.append(json_path)
    return results


def _emit_class_enum_bundles(excel_dir, schema_dir, schema_wins):
    """Merge __class__/__enum__ xlsx+json with schema bundles; write json only."""
    from XlsParser.XlsClass import CLASS_BASENAME, collect_class_defs_from_xlsx, collect_class_defs_from_json
    from XlsParser.XlsEnum import ENUM_BASENAME, collect_enum_defs_from_xlsx, collect_enum_defs_from_json

    schema_dir = os.path.abspath(schema_dir) if schema_dir else None
    excel_dir = os.path.abspath(excel_dir) if excel_dir else None
    if schema_dir:
        os.makedirs(schema_dir, exist_ok=True)
    if excel_dir:
        os.makedirs(excel_dir, exist_ok=True)

    class_schema_json = os.path.join(schema_dir, CLASS_JSON) if schema_dir else None
    enum_schema_json = os.path.join(schema_dir, ENUM_JSON) if schema_dir else None
    class_defs = _merge_type_defs(
        collect_class_defs_from_xlsx, collect_class_defs_from_json,
        CLASS_BASENAME, excel_dir, class_schema_json, schema_wins,
    )
    enum_defs = _merge_type_defs(
        collect_enum_defs_from_xlsx, collect_enum_defs_from_json,
        ENUM_BASENAME, excel_dir, enum_schema_json, schema_wins,
    )

    results = []
    class_json_list = _class_defs_to_json_list(class_defs)
    enum_json_list = _enum_defs_to_json_list(enum_defs)
    if class_json_list:
        json_path = class_schema_json if schema_dir else (
            os.path.join(excel_dir, CLASS_JSON) if excel_dir else None
        )
        results.extend(_write_type_bundle_json(class_json_list, json_path))
        if excel_dir and schema_dir:
            excel_json = os.path.join(excel_dir, CLASS_JSON)
            if os.path.abspath(excel_json) != os.path.abspath(json_path):
                results.extend(_write_type_bundle_json(class_json_list, excel_json))
        if schema_dir:
            _remove_kind0_schemas_for_workbook(schema_dir, CLASS_WORKBOOK)
    if enum_json_list:
        json_path = enum_schema_json if schema_dir else (
            os.path.join(excel_dir, ENUM_JSON) if excel_dir else None
        )
        results.extend(_write_type_bundle_json(enum_json_list, json_path))
        if excel_dir and schema_dir:
            excel_json = os.path.join(excel_dir, ENUM_JSON)
            if os.path.abspath(excel_json) != os.path.abspath(json_path):
                results.extend(_write_type_bundle_json(enum_json_list, excel_json))
        if schema_dir:
            _remove_kind0_schemas_for_workbook(schema_dir, ENUM_WORKBOOK)
    return results, class_defs, enum_defs


def _copy_type_bundles(schema_dir, dest_dir):
    """data->xlsx: merge type bundles into dest_dir/__class__.json and __enum__.json only."""
    from XlsParser.XlsClass import CLASS_BASENAME, collect_class_defs_from_xlsx, collect_class_defs_from_json
    from XlsParser.XlsEnum import ENUM_BASENAME, collect_enum_defs_from_xlsx, collect_enum_defs_from_json

    dest_dir = os.path.abspath(dest_dir)
    schema_dir = os.path.abspath(schema_dir)
    os.makedirs(dest_dir, exist_ok=True)

    class_schema_json = os.path.join(schema_dir, CLASS_JSON)
    enum_schema_json = os.path.join(schema_dir, ENUM_JSON)
    class_defs = _merge_type_defs(
        collect_class_defs_from_xlsx, collect_class_defs_from_json,
        CLASS_BASENAME, dest_dir, class_schema_json, schema_wins=True,
    )
    enum_defs = _merge_type_defs(
        collect_enum_defs_from_xlsx, collect_enum_defs_from_json,
        ENUM_BASENAME, dest_dir, enum_schema_json, schema_wins=True,
    )

    results = []
    class_json_list = _class_defs_to_json_list(class_defs)
    if class_json_list:
        results.extend(_write_type_bundle_json(class_json_list, os.path.join(dest_dir, CLASS_JSON)))
    enum_json_list = _enum_defs_to_json_list(enum_defs)
    if enum_json_list:
        results.extend(_write_type_bundle_json(enum_json_list, os.path.join(dest_dir, ENUM_JSON)))
    return results


EXCEL_DEFAULT_COL_WIDTH = 8.43
EXPORT_COL_WIDTH = EXCEL_DEFAULT_COL_WIDTH * 1.5

STRING_ALIASES = frozenset(
    {"string", "i18nstring", "lang", "json", "bit32", "bit64", "bit"}
)


def field_remarks(field: dict) -> str:
    if not field:
        return ""
    return field.get("remarks") or ""


def is_excel_text_type(typ: str) -> bool:
    t = (typ or "").strip().lower()
    if not t:
        return False
    if t in STRING_ALIASES or t == "string":
        return True
    if t.startswith("list<") or t.startswith("map<"):
        return True
    return False


def format_bit_mask_cell(val):
    """Convert stored bit mask int back to Excel [bitIndex,...] text."""
    if val is None or val == "":
        return "[]"
    if isinstance(val, str):
        return val
    if isinstance(val, list):
        return "[" + ",".join(str(x) for x in val) + "]"
    try:
        n = int(val)
    except Exception:
        return str(val)
    if n == -1:
        return "[-1]"
    bits = []
    i = 0
    while n:
        if n & 1:
            bits.append(str(i))
        n >>= 1
        i += 1
    return "[" + ",".join(bits) + "]"


def _class_field_names(class_name: str):
    typ = Type.get(class_name)
    if typ is not None and typ.isClass() and typ.fields:
        return [f.name for f in typ.fields]
    return None


def encode_value_for_excel(typ: str, val):
    """Encode json/runtime value into Excel cell text/number for re-import."""
    if val is None:
        return None
    if isinstance(val, bool):
        return 1 if val else 0
    t = (typ or "").strip()
    tl = t.lower()
    if tl in ("bit32", "bit64", "bit"):
        return format_bit_mask_cell(val)
    if tl in ("string", "i18nstring", "lang"):
        return str(val)
    if t.startswith("list<") and t.endswith(">"):
        inner = t[5:-1].strip()
        if not isinstance(val, list):
            return str(val)
        class_fields = _class_field_names(inner)
        if class_fields:
            encoded = []
            for elem in val:
                if isinstance(elem, dict):
                    encoded.append([elem.get(k) for k in class_fields])
                else:
                    encoded.append(elem)
            return json.dumps(encoded, ensure_ascii=False)
        # nested list / primitives
        return json.dumps(val, ensure_ascii=False)
    if t.startswith("map<") and t.endswith(">"):
        if isinstance(val, dict):
            return json.dumps(val, ensure_ascii=False)
        return val
    class_fields = _class_field_names(t)
    if class_fields and isinstance(val, dict):
        return format_bracket_list([val.get(k) for k in class_fields])
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False)
    return val


def format_int_with_base(val, int_base):
    if val is None or isinstance(val, bool):
        return val
    try:
        n = int(val)
    except Exception:
        return val
    if int_base == 16:
        return hex(n)
    if int_base == 2:
        return bin(n)
    return n


def format_cell_value(typ: str, val, int_base=None):
    if val is None or val == "":
        t = (typ or "").strip()
        if t.startswith("list<"):
            return "[]"
        if t.startswith("map<"):
            return "{}"
        return "" if is_excel_text_type(typ) else None
    if int_base in (16, 2) and not isinstance(val, (list, dict, bool)):
        return format_int_with_base(val, int_base)
    return encode_value_for_excel(typ, val)


def format_leaf_value(val, int_base=None):
    if val is None:
        return None
    if isinstance(val, bool):
        return 1 if val else 0
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False)
    if int_base in (16, 2):
        return format_int_with_base(val, int_base)
    return val


def format_bracket_list(values) -> str:
    parts = []
    for v in values:
        if v is None:
            parts.append("nil")
        elif isinstance(v, bool):
            parts.append("1" if v else "0")
        elif isinstance(v, (list, dict)):
            parts.append(json.dumps(v, ensure_ascii=False))
        else:
            parts.append(str(v))
    return "[" + ",".join(parts) + "]"


def format_split_element_cell(field_type: str, layout: dict, elem):
    """Format one list element for P2 (one element per column)."""
    if elem is None:
        return None
    if isinstance(elem, dict):
        element_keys = layout.get("elementKeys")
        if not element_keys:
            typ = Type.get((field_type or "").strip())
            # list<Class> → Class
            inner = field_type or ""
            if inner.startswith("list<") and inner.endswith(">"):
                inner = inner[5:-1].strip()
                typ = Type.get(inner)
            if typ is not None and typ.isClass() and typ.fields:
                element_keys = [f.name for f in typ.fields]
            else:
                element_keys = list(elem.keys())
        return format_bracket_list([elem.get(k) for k in element_keys])
    if isinstance(elem, (list, tuple)):
        return format_bracket_list(list(elem))
    return format_leaf_value(elem)


def field_layout_raw(field: dict) -> dict:
    layout = field.get("layout") if field else None
    return layout if isinstance(layout, dict) else {}


def field_layout(field: dict) -> dict:
    layout = field_layout_raw(field)
    if layout.get("mode") == "split":
        return layout
    return {}


def field_int_base(field: dict):
    return field_layout_raw(field).get("intBase")


def field_split_width(field: dict, sample_value=None) -> int:
    layout = field_layout(field)
    if not layout:
        return 1
    col_span = int(layout.get("colSpan") or 1)
    keys = layout.get("keys") or []
    needed = col_span
    if keys:
        element_span = int(layout.get("elementSpan") or len(keys) or 1)
        if isinstance(sample_value, list):
            needed = max(col_span, len(sample_value) * element_span)
        elif isinstance(sample_value, dict):
            needed = max(col_span, element_span)
    else:
        if isinstance(sample_value, list):
            needed = max(col_span, len(sample_value))
    return max(1, needed)


def expand_split_constraint_cells(field: dict, width: int) -> list:
    layout = field_layout(field)
    keys = layout.get("keys") or []
    if not keys:
        return [""] * width
    element_span = int(layout.get("elementSpan") or len(keys) or 1)
    cells = []
    for i in range(width):
        key = keys[i % element_span]
        cells.append(".%s" % key)
    return cells


def expand_split_data_cells(field: dict, value, width: int) -> list:
    layout = field_layout(field)
    field_type = field.get("type") or "int32"
    int_base = field_int_base(field)
    if not layout:
        return [format_cell_value(field_type, value, int_base)]
    keys = layout.get("keys") or []
    cells = [None] * width
    if keys:
        element_span = int(layout.get("elementSpan") or len(keys) or 1)
        if isinstance(value, dict):
            for i, key in enumerate(keys):
                if i < width:
                    cells[i] = format_leaf_value(value.get(key), int_base)
        elif isinstance(value, list):
            for elem_idx, elem in enumerate(value):
                base = elem_idx * element_span
                if isinstance(elem, dict):
                    for i, key in enumerate(keys):
                        pos = base + i
                        if pos < width:
                            cells[pos] = format_leaf_value(elem.get(key), int_base)
                elif base < width:
                    cells[base] = format_leaf_value(elem, int_base)
        elif value is not None and width > 0:
            cells[0] = format_leaf_value(value, int_base)
    else:
        if isinstance(value, list):
            for i, elem in enumerate(value):
                if i < width:
                    cells[i] = format_split_element_cell(field_type, layout, elem)
        elif value is not None and width > 0:
            cells[0] = format_cell_value(field_type, value, int_base)
    return cells


def apply_text_number_format(cell) -> None:
    if cell.value is None:
        cell.value = ""
    elif not isinstance(cell.value, str):
        cell.value = str(cell.value)
    cell.number_format = "@"


def apply_export_column_widths(ws, max_col: int) -> None:
    from openpyxl.utils import get_column_letter

    for c in range(1, max(1, max_col) + 1):
        ws.column_dimensions[get_column_letter(c)].width = EXPORT_COL_WIDTH


def paint_header_black(ws, max_row: int, max_col: int) -> None:
    fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
    font = Font(color="FFFFFFFF", bold=True)
    thin = Side(style="thin", color="FF9E9E9E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            # Keep None as None (empty "" breaks constraint-row parsing).
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def save_workbook(wb, xlsx_out: Path) -> None:
    tmp = xlsx_out.with_name(xlsx_out.stem + ".__export_tmp__.xlsx")
    try:
        wb.save(tmp)
        try:
            if xlsx_out.exists():
                xlsx_out.unlink()
            tmp.replace(xlsx_out)
        except OSError as e:
            raise SystemExit(
                "无法写入导出文件（可能正被 Excel/WPS 打开）: %s\n请关闭后重试。\n%s"
                % (xlsx_out, e)
            ) from e
    except PermissionError as e:
        raise SystemExit(
            "无法写入导出文件（可能正被 Excel/WPS 打开）: %s\n请关闭后重试。\n%s"
            % (xlsx_out, e)
        ) from e
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass


def keyed_schema_fields(fields: list) -> list:
    """Ensure id is first for 2D tables."""
    user = []
    id_field = None
    for f in fields or []:
        n = (f.get("name") or "").strip()
        if not n:
            continue
        if n.lower() == "id":
            id_field = dict(f)
            id_field["name"] = "id"
        else:
            user.append(f)
    if id_field is None:
        id_field = {"name": "id", "type": "int32", "displayName": "id"}
    return [id_field] + user


def default_xlsx_path(schema: dict, out_dir: Path) -> Path:
    workbook = schema.get("workbook")
    if workbook:
        return Path(out_dir) / os.path.basename(workbook)
    name = schema.get("name") or "table"
    sheet_name = schema.get("sheetName") or "data"
    # Fallback when workbook meta missing: stem from logical name.
    if sheet_name != "data" and name.endswith("_" + sheet_name):
        stem = name[: -(len(sheet_name) + 1)]
    else:
        stem = name
    return Path(out_dir) / ("%s.xlsx" % stem)


def sheet_title_from_schema(schema: dict) -> str:
    """Excel sheet tab title.

    - ``sheetName == "data"`` or no displayName → title is just ``sheetName``
      (workbook already carries display via ``name@中文.xlsx``, avoid ``data@物品``).
    - Else (multi-sheet non-data) → ``SheetName@displayName``.
    """
    sheet_name = (schema.get("sheetName") or "data").strip() or "data"
    disp = (schema.get("displayName") or schema.get("comment") or "").strip()
    if sheet_name == "data" or not disp:
        title = sheet_name
    else:
        title = "%s@%s" % (sheet_name, disp)
    # Excel sheet title limits
    title = re.sub(r'[:\\/?*\[\]]', "_", title)
    return title[:31] or "data"


def is_enum_schema(schema) -> bool:
    """Enum schemas are kind=0 with enumType."""
    return isinstance(schema, dict) and bool(schema.get("enumType"))


def is_class_schema(schema) -> bool:
    if not isinstance(schema, dict):
        return False
    if is_enum_schema(schema):
        return False
    kind = schema.get("kind")
    if kind == KIND_CLASS:
        return True
    if kind != KIND_1D:
        return False
    workbook = os.path.basename(schema.get("workbook") or "")
    if workbook == CLASS_WORKBOOK:
        return True
    # Class schemas: name == typename (tables usually differ, e.g. entity_Hero vs Entity_Hero)
    name = schema.get("name") or ""
    type_name = schema.get("typename") or ""
    return bool(name) and name == type_name


def fill_worksheet_from_schema(ws, schema: dict, data) -> None:
    """Write one sheet's header + rows into an existing worksheet."""
    fields = schema.get("fields") or []
    kind = schema.get("kind", KIND_1D)

    if kind == KIND_1D:
        header = ["##key", "type", "value", "desc", "tags", "##end"]
        ws.append(header)
        obj = data if isinstance(data, dict) else (
            data[0] if isinstance(data, list) and data else {}
        )
        data_row = 1
        for f in fields:
            n = f.get("name") or ""
            if not n:
                continue
            typ = f.get("type") or "int32"
            data_row += 1
            cell_val = format_cell_value(
                typ,
                obj.get(n) if isinstance(obj, dict) else None,
                field_int_base(f),
            )
            ws.append(
                [
                    n,
                    typ,
                    cell_val,
                    f.get("displayName") or f.get("comment") or n,
                    format_tags_cell(f.get("tags"), f.get("group")),
                    "",
                ]
            )
            apply_text_number_format(ws.cell(data_row, 1))
            apply_text_number_format(ws.cell(data_row, 2))
            apply_text_number_format(ws.cell(data_row, 4))
            if is_excel_text_type(typ) or field_int_base(f) in (16, 2):
                apply_text_number_format(ws.cell(data_row, 3))
            rem = field_remarks(f)
            if rem:
                ws.cell(data_row, 4).comment = Comment(rem, "xls2cfg")
        paint_header_black(ws, 1, len(header))
        apply_export_column_widths(ws, len(header))
    else:
        fields = keyed_schema_fields(fields)
        named = [f for f in fields if f.get("name")]
        rows = [r for r in (data if isinstance(data, list) else []) if isinstance(r, dict)]

        # Expand each logical field into 1..N Excel columns (layout.colSpan,
        # grown if any data row needs more slots).
        col_specs = []  # (field, width, start_col_1based)
        col = 1
        for f in named:
            fname = f.get("name")
            width = field_split_width(f, None)
            for r in rows:
                width = max(width, field_split_width(f, r.get(fname)))
            col_specs.append((f, width, col))
            col += width
        end_col = col  # includes trailing ##end column index

        remarks_row = []
        names_row = []
        types_row = []
        constraints_row = []
        tags_row = []
        for f, width, _start in col_specs:
            remarks_row.append(f.get("displayName") or f.get("comment") or f.get("name") or "")
            remarks_row.extend([""] * (width - 1))
            names_row.append(f.get("name"))
            names_row.extend([None] * (width - 1))
            types_row.append(f.get("type") or "int32")
            types_row.extend([None] * (width - 1))
            constraints_row.extend(expand_split_constraint_cells(f, width))
            # Non-split fields keep their constraint text on the first column.
            if width == 1 and not field_layout(f):
                constraints_row[-1] = f.get("constraint") or ""
            tags_row.append(format_tags_cell(f.get("tags"), f.get("group")))
            tags_row.extend([""] * (width - 1))

        remarks_row.append("##end")
        names_row.append("")
        types_row.append("")
        constraints_row.append("")
        tags_row.append("")

        ws.append(remarks_row)
        ws.append(names_row)
        ws.append(types_row)
        ws.append(constraints_row)
        ws.append(tags_row)

        paint_header_black(ws, 5, end_col)
        apply_export_column_widths(ws, end_col)
        for f, width, start_col in col_specs:
            rem = field_remarks(f)
            if rem:
                ws.cell(1, start_col).comment = Comment(rem, "xls2cfg")

        # Merge header cells (rows 1-3) after styling (MergedCell is read-only).
        for f, width, start_col in col_specs:
            if width <= 1:
                continue
            end = start_col + width - 1
            for row_idx in (1, 2, 3):
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=start_col,
                    end_row=row_idx,
                    end_column=end,
                )

        text_cols = set()
        for f, width, start_col in col_specs:
            if field_layout(f) or is_excel_text_type(f.get("type") or ""):
                for c in range(start_col, start_col + width):
                    text_cols.add(c)

        for r in rows:
            line = []
            for f, width, _start in col_specs:
                cells = expand_split_data_cells(f, r.get(f.get("name")), width)
                line.extend(cells)
            line.append("")
            ws.append(line)
            row_idx = ws.max_row
            for c in text_cols:
                cell = ws.cell(row_idx, c)
                if isinstance(cell.value, str) or cell.value is None:
                    apply_text_number_format(cell)


def write_excel_from_schema_json(schema: dict, data, xlsx_out: Path) -> Path:
    if openpyxl is None:
        raise SystemExit("openpyxl required: pip install openpyxl")
    if is_enum_schema(schema):
        return write_enum_workbook([schema], Path(xlsx_out))
    if is_class_schema(schema):
        return write_class_workbook([schema], Path(xlsx_out))

    xlsx_out = Path(xlsx_out)
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    if xlsx_out.suffix.lower() != ".xlsx":
        xlsx_out = xlsx_out.with_suffix(".xlsx")
    # Rewrite default stem (schema name) to workbook filename when meta present.
    name = schema.get("name") or "table"
    if xlsx_out.stem == name and schema.get("workbook"):
        xlsx_out = default_xlsx_path(schema, xlsx_out.parent)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title_from_schema(schema)
    fill_worksheet_from_schema(ws, schema, data)
    save_workbook(wb, xlsx_out)
    return xlsx_out


def write_multi_sheet_workbook(sheet_items, xlsx_out: Path) -> Path:
    """sheet_items: [(sheetIndex, schema, data), ...] sorted by sheetIndex."""
    if openpyxl is None:
        raise SystemExit("openpyxl required: pip install openpyxl")
    xlsx_out = Path(xlsx_out)
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    first = True
    for _idx, schema, data in sheet_items:
        title = sheet_title_from_schema(schema)
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title)
        fill_worksheet_from_schema(ws, schema, data)
    save_workbook(wb, xlsx_out)
    return xlsx_out


def write_class_workbook(class_schemas, xlsx_out: Path) -> Path:
    """Restore __class__.xlsx from kind=0 class schemas (fields as list<__Field__>)."""
    if openpyxl is None:
        raise SystemExit("openpyxl required: pip install openpyxl")
    xlsx_out = Path(xlsx_out)
    if xlsx_out.is_dir() or xlsx_out.suffix.lower() != ".xlsx":
        xlsx_out = Path(xlsx_out) / CLASS_WORKBOOK
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)

    schemas = list(class_schemas)
    max_fields = 1
    for s in schemas:
        max_fields = max(max_fields, len(s.get("fields") or []))
    field_width = max_fields * 4

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"

    remarks = ["类型名", "备注名", "字段定义"] + [None] * (field_width - 1) + ["##end"]
    names = ["typename", "comment", "fields"] + [None] * (field_width - 1)
    types = ["string", "string", "list<__Field__>"] + [None] * (field_width - 1)
    constraints = [None, None]
    for i in range(field_width):
        constraints.append([".name", ".type", ".comment", ".tags"][i % 4])
    tags = [None] * (2 + field_width)

    ws.append(remarks)
    ws.append(names)
    ws.append(types)
    ws.append(constraints)
    ws.append(tags)

    fields_end = 2 + field_width
    end_col = fields_end + 1  # includes ##end
    # Style before merge (MergedCell values are read-only).
    paint_header_black(ws, 5, end_col)
    apply_export_column_widths(ws, end_col)
    if field_width > 1:
        for row_idx in (1, 2, 3):
            ws.merge_cells(
                start_row=row_idx,
                start_column=3,
                end_row=row_idx,
                end_column=fields_end,
            )

    for schema in schemas:
        fields = [f for f in (schema.get("fields") or []) if f.get("name")]
        line = [
            schema.get("name") or schema.get("typename") or "",
            schema.get("displayName") or schema.get("comment") or "",
        ]
        for f in fields:
            tags_val = format_tags_cell(f.get("tags"), f.get("group"))
            line.extend([
                f.get("name") or "",
                f.get("type") or "int32",
                f.get("displayName") or f.get("comment") or "",
                tags_val if tags_val else None,
            ])
        while len(line) < fields_end:
            line.append(None)
        line.append(None)  # ##end column
        ws.append(line[:end_col])

    save_workbook(wb, xlsx_out)
    return xlsx_out


def write_enum_workbook(enum_schemas, xlsx_out: Path) -> Path:
    """Restore __enum__.xlsx from enum schemas (kind=0 + enumType, fields as list<__EnumField__>)."""
    if openpyxl is None:
        raise SystemExit("openpyxl required: pip install openpyxl")
    xlsx_out = Path(xlsx_out)
    if xlsx_out.is_dir() or xlsx_out.suffix.lower() != ".xlsx":
        xlsx_out = Path(xlsx_out) / ENUM_WORKBOOK
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)

    schemas = sorted(enum_schemas, key=lambda s: s.get("name") or "")
    max_fields = 1
    for s in schemas:
        max_fields = max(max_fields, len(s.get("fields") or []))
    field_width = max_fields * 4

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"

    remarks = ["类型名", "备注名", "底层类型", "可组合", "导出类型", "字段定义"] + [None] * (field_width - 1) + ["##end"]
    names = ["typename", "comment", "enumType", "flags", "exportType", "fields"] + [None] * (field_width - 1)
    types = ["string", "string", "string", "bool", "bool", "list<__EnumField__>"] + [None] * (field_width - 1)
    constraints = [None, None, None, None, None]
    for i in range(field_width):
        constraints.append([".name", ".value", ".comment", ".tags"][i % 4])
    tags = [None] * (5 + field_width)

    ws.append(remarks)
    ws.append(names)
    ws.append(types)
    ws.append(constraints)
    ws.append(tags)

    fields_end = 5 + field_width
    end_col = fields_end + 1  # includes ##end
    paint_header_black(ws, 5, end_col)
    apply_export_column_widths(ws, end_col)
    if field_width > 1:
        for row_idx in (1, 2, 3):
            ws.merge_cells(
                start_row=row_idx,
                start_column=6,
                end_row=row_idx,
                end_column=fields_end,
            )

    for schema in schemas:
        items = [f for f in (schema.get("fields") or []) if f.get("name")]
        line = [
            schema.get("name") or schema.get("typename") or "",
            schema.get("displayName") or schema.get("comment") or "",
            schema.get("enumType") or "int32",
            1 if schema.get("flags") else 0,
            1 if schema.get("exportType") else 0,
        ]
        for f in items:
            tags_val = format_tags_cell(f.get("tags"), f.get("group"))
            line.extend([
                f.get("name") or "",
                f.get("value") if f.get("value") is not None else None,
                f.get("displayName") or f.get("comment") or "",
                tags_val if tags_val else None,
            ])
        while len(line) < fields_end:
            line.append(None)
        line.append(None)  # ##end column
        ws.append(line[:end_col])

    save_workbook(wb, xlsx_out)
    return xlsx_out


# --- export: schema + json/binary -> xlsx ---

def schema_binary_fields(schema: dict, tags=None) -> list:
    prev_tags = Config.tags
    if tags is not None:
        Config.tags = tags
    try:
        fields = []
        for f in schema.get("fields") or []:
            name = f.get("name")
            if not name:
                continue
            field_tags_str = f.get("tags") or ""
            field_tags = (
                [t.strip() for t in field_tags_str.split(",") if t.strip()]
                if field_tags_str else None
            )
            if not Config.isNeedExportTags(field_tags):
                continue
            fields.append((name, Type.getOrCreate(f.get("type") or "int32")))
        return fields
    finally:
        Config.tags = prev_tags


def read_binary_data(schema: dict, binary_path: Path, tags=None):
    field_types = schema_binary_fields(schema, tags=tags)
    bs = ByteStream()
    bs.ReadFile(str(binary_path))
    if schema.get("kind", KIND_1D) == KIND_1D:
        row = {}
        for name, typ in field_types:
            row[name] = bs.ReadValue(typ)
        return row
    data_row = bs.ReadUInt16()
    rows = []
    for _ in range(data_row):
        row = {}
        for name, typ in field_types:
            row[name] = bs.ReadValue(typ)
        rows.append(row)
    return rows


def load_export_data(schema: dict, json_path=None, binary_path=None, tags=None):
    json_path = Path(json_path) if json_path else None
    binary_path = Path(binary_path) if binary_path else None
    if json_path and json_path.exists():
        return json.loads(json_path.read_text(encoding="utf-8"))
    if binary_path and binary_path.exists():
        return read_binary_data(schema, binary_path, tags=tags)
    return {} if schema.get("kind", KIND_1D) == KIND_1D else []


def register_class_schemas_from_dir(schema_dir: Path) -> None:
    """Load __class__.json (and leftover kind=0 files) into Type registry."""
    from XlsParser.XlsClass import collect_class_defs_from_json, register_class_defs

    schema_dir = Path(schema_dir)
    if not schema_dir.is_dir():
        return
    bundle = schema_dir / CLASS_JSON
    if bundle.is_file():
        register_class_defs(collect_class_defs_from_json(str(bundle), apply_tags_filter=False))
    for path in sorted(schema_dir.glob("*.json")):
        if path.name in TYPE_BUNDLE_FILES:
            continue
        try:
            obj = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not is_class_schema(obj):
            continue
        class_name = obj.get("typename") or ""
        if not class_name:
            continue
        existing = Type.get(class_name)
        if isinstance(existing, Type) and existing.isClass() and existing.fields:
            continue
        fields = []
        for f in obj.get("fields") or []:
            n = f.get("name")
            if not n:
                continue
            fields.append({
                "type": f.get("type") or "int32",
                "name": n,
                "comment": f.get("displayName") or f.get("comment") or "",
                "tags": None,
                "remarks": f.get("remarks") or "",
                "group": f.get("group"),
            })
        if not fields:
            continue
        type_comment = obj.get("displayName") or obj.get("comment") or ""
        if isinstance(existing, Type):
            existing.fields = []
            existing.idFieldIdx = -1
            for f in fields:
                existing.defineField(
                    f["type"], f["name"], comment=f["comment"], remarks=f["remarks"],
                    group=f.get("group"),
                )
            if type_comment:
                existing.comment = type_comment
        else:
            typ = Type.createClass(class_name, fields)
            if type_comment:
                typ.comment = type_comment


def register_enum_schemas_from_dir(schema_dir: Path) -> None:
    """Load __enum__.json (and leftover kind=0 enum files) into Type registry."""
    from XlsParser.XlsEnum import collect_enum_defs_from_json, register_enum_defs

    schema_dir = Path(schema_dir)
    if not schema_dir.is_dir():
        return
    bundle = schema_dir / ENUM_JSON
    if bundle.is_file():
        register_enum_defs(collect_enum_defs_from_json(str(bundle), apply_tags_filter=False))
    for path in sorted(schema_dir.glob("*.json")):
        if path.name in TYPE_BUNDLE_FILES:
            continue
        try:
            obj = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not is_enum_schema(obj):
            continue
        enum_name = obj.get("typename") or ""
        if not enum_name:
            continue
        existing = Type.get(enum_name)
        if isinstance(existing, Type) and existing.isEnum():
            continue
        items = []
        for f in obj.get("fields") or []:
            n = f.get("name")
            if not n:
                continue
            items.append({
                "name": n,
                "value": f.get("value"),
                "comment": f.get("displayName") or f.get("comment") or "",
                "tags": None,
            })
        Type.createEnum(
            enum_name,
            enumType=obj.get("enumType") or "int32",
            items=items,
            comment=obj.get("displayName") or obj.get("comment") or None,
            flags=bool(obj.get("flags")),
            exportType=bool(obj.get("exportType")),
        )


def write_xlsx_from_data(
    schema_path: Path,
    out_xlsx_dir: Path,
    json_path=None,
    binary_path=None,
    tags=None,
) -> Path:
    """Single table: schema + json|binary -> {out_xlsx_dir}/{schema.workbook}."""
    schema_path = Path(schema_path)
    out_xlsx_dir = Path(out_xlsx_dir)
    out_xlsx_dir.mkdir(parents=True, exist_ok=True)
    register_enum_schemas_from_dir(schema_path.parent)
    register_class_schemas_from_dir(schema_path.parent)
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    data = load_export_data(
        schema, json_path=json_path, binary_path=binary_path, tags=tags
    )
    xlsx_out = default_xlsx_path(schema, out_xlsx_dir)
    path = write_excel_from_schema_json(schema, data, xlsx_out)
    _copy_type_bundles(schema_path.parent, out_xlsx_dir)
    return path


def write_xlsx_from_dir(
    schema_dir: Path,
    out_xlsx_dir: Path,
    json_dir=None,
    binary_dir=None,
    tags=None,
) -> list:
    """Batch: schema_dir + json_dir|binary_dir -> out_xlsx_dir.

    Tables with the same workbook are merged into one multi-sheet xlsx (sheetIndex order).
    __class__.json / __enum__.json are copied as-is (not rebuilt as xlsx).
    """
    schema_dir = Path(schema_dir)
    json_dir = Path(json_dir) if json_dir else None
    binary_dir = Path(binary_dir) if binary_dir else None
    out_xlsx_dir = Path(out_xlsx_dir)
    out_xlsx_dir.mkdir(parents=True, exist_ok=True)
    results = []
    if not schema_dir.is_dir():
        raise SystemExit("schema dir not found: %s" % schema_dir)
    if not json_dir and not binary_dir:
        raise SystemExit("--from-json-dir or --from-binary-dir required")

    register_enum_schemas_from_dir(schema_dir)
    register_class_schemas_from_dir(schema_dir)
    groups = {}  # workbook filename -> [(sheetIndex, name, schema, data)]

    for schema_path in sorted(schema_dir.glob("*.json")):
        if schema_path.name in TYPE_BUNDLE_FILES:
            continue
        schema = json.loads(schema_path.read_text(encoding="utf-8"))
        if not isinstance(schema, dict):
            continue
        if is_enum_schema(schema) or is_class_schema(schema):
            continue
        name = schema_path.stem
        json_path = (json_dir / (name + ".json")) if json_dir else None
        binary_path = (binary_dir / (name + ".bytes")) if binary_dir else None
        data = load_export_data(
            schema,
            json_path=json_path if json_path and json_path.exists() else None,
            binary_path=binary_path if binary_path and binary_path.exists() else None,
            tags=tags,
        )
        wb_name = os.path.basename(
            schema.get("workbook") or default_xlsx_path(schema, out_xlsx_dir).name
        )
        sheet_index = schema.get("sheetIndex", 0)
        groups.setdefault(wb_name, []).append((sheet_index, name, schema, data))

    for wb_name, items in sorted(groups.items()):
        items.sort(key=lambda x: (x[0], x[1]))
        out = out_xlsx_dir / wb_name
        path = write_multi_sheet_workbook(
            [(idx, schema, data) for idx, _name, schema, data in items],
            out,
        )
        results.append(str(path))
        print("export", path)

    results.extend(_copy_type_bundles(schema_dir, out_xlsx_dir))
    return results


# ---- xlsx -> schema + json ----

def _top_level_comma(s):
    depth = 0
    for i, ch in enumerate(s or ""):
        if ch == "<":
            depth += 1
        elif ch == ">":
            depth -= 1
        elif ch == "," and depth == 0:
            return i
    return -1


def iter_type_atoms(type_str):
    t = (type_str or "").strip()
    if not t:
        return
    if t.startswith("list<") and t.endswith(">"):
        for a in iter_type_atoms(t[5:-1].strip()):
            yield a
        return
    if t.startswith("map<") and t.endswith(">"):
        inner = t[4:-1]
        comma = _top_level_comma(inner)
        if comma < 0:
            yield t
            return
        for a in iter_type_atoms(inner[:comma].strip()):
            yield a
        for a in iter_type_atoms(inner[comma + 1 :].strip()):
            yield a
        return
    yield t


def collect_needed_class_names(type_strings, class_names_from_file):
    """Closure of __class__ type names referenced by type_strings."""
    class_set = set(class_names_from_file or [])
    needed = set()

    def add_deps(name):
        if name not in class_set or name in needed:
            return
        needed.add(name)
        typ = Type.get(name)
        if typ is None or not typ.isClass() or not typ.fields:
            return
        for f in typ.fields:
            ft = f.type.fullTypename if f.type else ""
            for atom in iter_type_atoms(ft):
                add_deps(atom)

    for ts in type_strings or []:
        for atom in iter_type_atoms(ts):
            add_deps(atom)
    return needed


def write_class_schema(typ, schema_dir, json_dir=None, sheet_index=0):
    """Write kind=0 class schema. Filename = type name.

    Does not write json data (kind=0 is schema/ClassDecl only).
    ``json_dir`` is accepted for call-site compatibility but ignored.
    """
    if typ is None or not typ.isClass() or typ.typename in ("__Field__", "__EnumField__"):
        return ""
    schema = typ.to_schema(name=typ.typename, kind=KIND_CLASS, class_name=typ.typename)
    out = {
        "kind": KIND_CLASS,
        "name": typ.typename,
    }
    if schema.get("displayName"):
        out["displayName"] = schema["displayName"]
    out["typename"] = typ.typename
    out["workbook"] = CLASS_WORKBOOK
    out["sheetName"] = "data"
    out["sheetIndex"] = sheet_index
    out["fields"] = schema["fields"]
    return _write_class_schema_dict(out, schema_dir, json_dir=None)


def write_enum_schema(typ, schema_dir, sheet_index=0):
    """Write enum schema (kind=0 + enumType). Filename = type name."""
    if typ is None or not typ.isEnum():
        return ""
    schema = typ.to_schema(name=typ.typename, kind=KIND_CLASS, class_name=typ.typename)
    out = {
        "kind": KIND_CLASS,
        "name": typ.typename,
    }
    if schema.get("displayName"):
        out["displayName"] = schema["displayName"]
    out["typename"] = typ.typename
    out["enumType"] = schema.get("enumType") or "int32"
    if schema.get("flags"):
        out["flags"] = True
    if schema.get("exportType"):
        out["exportType"] = True
    out["workbook"] = ENUM_WORKBOOK
    out["sheetName"] = "data"
    out["sheetIndex"] = sheet_index
    out["fields"] = schema["fields"]
    return _write_class_schema_dict(out, schema_dir, json_dir=None)


def _write_class_schema_dict(out, schema_dir, json_dir=None):
    os.makedirs(schema_dir, exist_ok=True)
    name = out.get("name") or ""
    schema_path = os.path.join(schema_dir, name + ".json")
    data = json.dumps(out, ensure_ascii=False, indent=2)
    with open(schema_path, "w", encoding="utf-8", newline="\n") as fd:
        fd.write(data)
        if not data.endswith("\n"):
            fd.write("\n")
    # kind=0: never write table/json data
    print("write", schema_dir, name)
    return schema_path


def write_class_schemas_from_excel(excel_dir, schema_dir, class_path=None):
    """Write unified __class__.json from xlsx+json (json overrides)."""
    from XlsParser.XlsClass import collect_class_defs

    base = class_path if class_path is not None else excel_dir
    os.makedirs(schema_dir, exist_ok=True)
    out_path = os.path.join(schema_dir, CLASS_JSON)
    prev_tags = Config.tags
    Config.tags = []
    try:
        defs = collect_class_defs(base, apply_tags_filter=False)
        if not defs:
            return []
        _write_json_list(out_path, _class_defs_to_json_list(defs))
        written = list(defs.keys())
    finally:
        Config.tags = prev_tags
    _remove_kind0_schemas_for_workbook(schema_dir, CLASS_WORKBOOK)
    return written


def write_enum_schemas_from_excel(excel_dir, schema_dir, enum_path=None):
    """Write unified __enum__.json from xlsx+json (json overrides)."""
    from XlsParser.XlsEnum import collect_enum_defs

    base = enum_path if enum_path is not None else excel_dir
    os.makedirs(schema_dir, exist_ok=True)
    out_path = os.path.join(schema_dir, ENUM_JSON)
    prev_tags = Config.tags
    Config.tags = []
    try:
        defs = collect_enum_defs(base, apply_tags_filter=False)
        if not defs:
            return []
        _write_json_list(out_path, _enum_defs_to_json_list(defs))
        written = list(defs.keys())
    finally:
        Config.tags = prev_tags
    _remove_kind0_schemas_for_workbook(schema_dir, ENUM_WORKBOOK)
    return written


def write_data_from_xlsx(
    xlsx_path,
    output_dir,
    output_formats,
    parser_for_format,
):
    """Write one workbook to {output}/{format}/ for every requested format.

    All importable sheets are written. Nested/dep types from __class__/__enum__
    are registered so codegen compiles, but are not written unless this xlsx
    *is* __class__/__enum__.
    """
    if load_workbook is None:
        raise SystemExit("openpyxl required: pip install openpyxl")

    from XlsParser.XlsParser import XlsParser

    formats = [f.strip() for f in (output_formats or []) if f and str(f).strip()]
    if not formats:
        raise SystemExit("output-formats required with --from-xlsx")

    xlsx_path = os.path.abspath(xlsx_path)
    output_dir = os.path.abspath(output_dir)
    excel_dir = os.path.dirname(xlsx_path)
    fileName = os.path.basename(xlsx_path)
    enum_names = readEnum(excel_dir) or []
    class_names = readClass(excel_dir) or []

    if fileName.startswith("__enum__") or fileName.startswith("__class__"):
        schema_dir = os.path.join(output_dir, "schema")
        os.makedirs(schema_dir, exist_ok=True)
        from XlsParser.XlsClass import register_class_defs
        from XlsParser.XlsEnum import register_enum_defs
        _results, class_defs, enum_defs = _emit_class_enum_bundles(
            excel_dir, schema_dir, schema_wins=False,
        )
        if fileName.startswith("__enum__"):
            register_enum_defs(enum_defs)
            defs = enum_defs
            bundle = os.path.join(schema_dir, ENUM_JSON)
            wb_name = ENUM_WORKBOOK
        else:
            register_class_defs(class_defs)
            defs = class_defs
            bundle = os.path.join(schema_dir, CLASS_JSON)
            wb_name = CLASS_WORKBOOK
        code_formats = [f for f in formats if f != "schema"]
        if code_formats:
            export_code_from_schema(bundle, output_dir, code_formats, parser_for_format)
        deps = [{
            "name": name,
            "kind": KIND_CLASS,
            "schema": bundle,
        } for name in defs]
        print(json.dumps({
            "ok": True,
            "name": wb_name,
            "tables": [],
            "deps": deps,
            "count": len(deps),
        }, ensure_ascii=False))
        return deps

    wb = load_workbook(filename=xlsx_path, data_only=True)
    loaded = []
    try:
        for idx, sn in enumerate(wb.sheetnames):
            if not is_importable_sheet_title(sn):
                continue
            loaded.append(Sheet(wb[sn], fileName, sn, idx))
    finally:
        wb.close()
    if not loaded:
        raise SystemExit("workbook has no importable sheets")

    XlsParser.ignoreGenTables = True
    tables = []
    type_strings = []
    sheet_typenames = []
    for fmt in formats:
        Parser = parser_for_format(fmt)
        if Parser is None:
            raise Exception("unknown outputFormat: %s" % fmt)
        out = os.path.join(output_dir, fmt)
        os.makedirs(out, exist_ok=True)
        fmt_types = []
        for sheet in loaded:
            parser = Parser(sheet, out)
            parser.parse()
            typ = getattr(parser, "type", None)
            if isinstance(typ, Type) and typ.typename not in fmt_types:
                fmt_types.append(typ.typename)
        if "writeClass" in Parser.__dict__:
            for name in fmt_types:
                typ = Type.get(name)
                if isinstance(typ, Type) and typ.isClass():
                    Parser.writeClass(typ, out)
        if fmt == formats[0]:
            sheet_typenames = list(fmt_types)
            for sheet in loaded:
                tables.append({
                    "name": sheet.filename,
                    "kind": KIND_1D if sheet.singleton else KIND_2D,
                    "sheet": sheet.sheetName,
                    "sheetIndex": sheet.sheetIndex,
                    "workbook": sheet.xlsFilename,
                    "schema": os.path.join(output_dir, "schema", sheet.filename + ".json") if "schema" in formats else None,
                    "data": os.path.join(output_dir, "json", sheet.filename + ".json") if "json" in formats else None,
                })
                for col in range(0, sheet.maxCol):
                    typ = sheet.col2type.get(col)
                    if typ is not None:
                        type_strings.append(typ.fullTypename)

    deps = []
    if class_names:
        needed = collect_needed_class_names(type_strings, class_names)
        for name in sorted(needed):
            if name in sheet_typenames:
                continue
            deps.append({
                "name": name,
                "kind": KIND_CLASS,
                "schema": os.path.join(excel_dir, CLASS_JSON),
            })
    if enum_names:
        enum_set = set(enum_names)
        needed_enums = set()
        for ts in type_strings:
            for atom in iter_type_atoms(ts):
                if atom in enum_set:
                    needed_enums.add(atom)
        for name in sorted(needed_enums):
            deps.append({
                "name": name,
                "kind": KIND_CLASS,
                "schema": os.path.join(excel_dir, ENUM_JSON),
            })

    print(json.dumps({
        "ok": True,
        "name": tables[0]["name"],
        "tables": tables,
        "deps": deps,
        "count": len(tables),
    }, ensure_ascii=False))
    return tables


def _load_schema_file_objects(path: Path):
    from XlsParser.TypeDefPath import load_types_json_list

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    if not isinstance(data, dict):
        return []
    if path.name in TYPE_BUNDLE_FILES or isinstance(data.get("types"), list):
        return load_types_json_list(str(path))
    if data.get("typename") or data.get("fields") is not None or data.get("enumType"):
        return [data]
    return load_types_json_list(str(path))


def _collect_schema_type_objects(schema_path):
    path = Path(schema_path)
    if not path.exists():
        raise SystemExit("schema path not found: %s" % path)
    if path.is_file():
        return _load_schema_file_objects(path)
    objs = []
    for p in sorted(path.glob("*.json")):
        objs.extend(_load_schema_file_objects(p))
    return objs


def _schema_fields_to_type_fields(fields):
    out = []
    for f in fields or []:
        name = f.get("name")
        ftype = f.get("type")
        if not name or not ftype:
            continue
        out.append({
            "type": ftype,
            "name": name,
            "comment": f.get("displayName") or f.get("comment") or "",
            "tags": f.get("tags"),
            "remarks": f.get("remarks") or "",
            "group": f.get("group"),
        })
    return out


def _is_builtin_type_atom(atom):
    return atom in Type.basicTypes or atom in Type.containerTypes or atom == "json"


def export_code_from_schema(schema_path, output_dir, output_formats, parser_for_format):
    """Generate language types from schema dir or a single schema json.

    Dependent types are registered as stubs when needed, but are not written.
    Language is decided by ``output_formats`` (csharp/go have class/enum codegen).
    """
    from XlsParser.XlsParser import XlsParser

    objs = _collect_schema_type_objects(schema_path)
    if not objs:
        raise SystemExit("no schema types found: %s" % schema_path)

    enum_objs = [o for o in objs if is_enum_schema(o)]
    class_objs = [o for o in objs if not is_enum_schema(o)]
    export_names = []

    for o in enum_objs:
        name = o.get("typename")
        if not name:
            continue
        export_names.append(name)
        Type.unregister(name)
        items = []
        for f in o.get("fields") or []:
            n = f.get("name")
            if not n:
                continue
            items.append({
                "name": n,
                "value": f.get("value"),
                "comment": f.get("displayName") or f.get("comment") or "",
                "tags": f.get("tags"),
            })
        Type.createEnum(
            name,
            enumType=o.get("enumType") or "int32",
            items=items,
            comment=o.get("displayName") or o.get("comment") or None,
            flags=bool(o.get("flags")),
            exportType=bool(o.get("exportType")),
        )

    class_names = []
    for o in class_objs:
        name = o.get("typename")
        if not name:
            continue
        class_names.append(name)
        export_names.append(name)

    known = set(class_names)
    for name, typ in list(Type.types.items()):
        if isinstance(typ, Type) and typ.isEnum():
            known.add(name)

    for o in class_objs:
        for f in o.get("fields") or []:
            for atom in iter_type_atoms(f.get("type") or ""):
                if _is_builtin_type_atom(atom) or atom in known:
                    continue
                existing = Type.get(atom)
                if isinstance(existing, Type):
                    continue
                Type.createClass(atom)
                known.add(atom)

    for o in class_objs:
        name = o.get("typename")
        if not name:
            continue
        Type.unregister(name)
        typ = Type.createClass(name, _schema_fields_to_type_fields(o.get("fields")))
        typ.comment = o.get("displayName") or o.get("comment") or None
        if o.get("kind") == KIND_1D:
            typ.singleton = True
        if not getattr(typ, "singleton", False) and typ.fields:
            id_idx = 0
            for i, f in enumerate(typ.fields):
                if f.name == "id":
                    id_idx = i
                    break
            typ.setIdField(id_idx)

    XlsParser.ignoreGenTables = True
    written = []
    for fmt in output_formats:
        Parser = parser_for_format(fmt)
        if Parser is None:
            raise Exception("unknown outputFormat: %s" % fmt)
        can_write_class = "writeClass" in Parser.__dict__
        can_write_enum = "writeEnum" in Parser.__dict__
        if not can_write_class and not can_write_enum:
            print("skip format %s (no type codegen)" % fmt)
            continue
        out = os.path.join(output_dir, fmt)
        os.makedirs(out, exist_ok=True)
        for name in export_names:
            typ = Type.get(name)
            if not isinstance(typ, Type):
                continue
            if typ.isEnum():
                if not getattr(typ, "exportType", False):
                    print("skip enum %s (exportType=false)" % name)
                    continue
                if can_write_enum:
                    Parser.writeEnum(typ, out)
                    written.append("%s/%s" % (fmt, name))
                else:
                    print("skip enum %s for format %s" % (name, fmt))
                continue
            if typ.isClass() and can_write_class:
                Parser.writeClass(typ, out)
                written.append("%s/%s" % (fmt, name))
            else:
                print("skip type %s for format %s" % (name, fmt))
    print(json.dumps({
        "ok": True,
        "schema": str(schema_path),
        "written": written,
        "count": len(written),
    }, ensure_ascii=False))
    return written
