#coding: utf-8
#@author: sundream
#@email: linguanglianglgl@gmail.com
#@date: 2021-02-05
#@version: 0.1.0

import sys
import optparse
import os
import json
import subprocess
from XlsParser.XlsParser import XlsParser
from XlsParser.Xls2PyParser import Xls2PyParser
from XlsParser.Xls2LuaParser import Xls2LuaParser
from XlsParser.Xls2LuaCvsParser import Xls2LuaCvsParser
from XlsParser.Xls2JsonParser import Xls2JsonParser
from XlsParser.Xls2BinaryParser import Xls2BinaryParser
from XlsParser.Xls2CSharpParser import Xls2CSharpParser
from XlsParser.Xls2GoParser import Xls2GoParser
from XlsParser.Xls2SchemaParser import Xls2SchemaParser
from XlsParser.XlsImportExport import (
    export_one,
    export_dir,
    import_xlsx_to_dirs,
    write_class_schemas_from_excel,
)
from XlsParser.Sheet import Sheet, getSheets, is_importable_sheet_title
from XlsParser.I18NExport import readI18nFile,writeI18nFile
from XlsParser.XlsClass import readClass
from XlsParser.Config import Config
from openpyxl import load_workbook

# optparse has no bool type; use choice. Invalid values hint: choose from 'true', 'false', '1', '0'
BOOL_CHOICES = ("true", "false", "1", "0")
BOOL_OPTION = {"type": "choice", "choices": list(BOOL_CHOICES)}


def choice_to_bool(value):
    if value is None:
        return None
    return value in ("true", "1")

def apply_json_config(jsonConfig):
    if "genMeta" in jsonConfig:
        Config.genMeta = jsonConfig.get("genMeta")
    if "genMetaDetail" in jsonConfig:
        Config.genMetaDetail = jsonConfig.get("genMetaDetail")
    if "genMetaHeader" in jsonConfig:
        Config.genMetaHeader = jsonConfig.get("genMetaHeader")
    if "pretty" in jsonConfig:
        Config.pretty = jsonConfig.get("pretty")
    if "defaults" in jsonConfig:
        Config.defaults = jsonConfig.get("defaults")
    if "tags" in jsonConfig:
        Config.tags = jsonConfig.get("tags")
    if "classNameFirstUpper" in jsonConfig:
        Config.classNameFirstUpper = jsonConfig.get("classNameFirstUpper")
    if "fieldNameFirstUpper" in jsonConfig:
        Config.fieldNameFirstUpper = jsonConfig.get("fieldNameFirstUpper")
    if "constraintSeperator" in jsonConfig:
        Config.constraintSeperator = jsonConfig.get("constraintSeperator")
    if "indent" in jsonConfig:
        Config.indent = jsonConfig.get("indent")
    if "localize" in jsonConfig:
        Config.localize = jsonConfig.get("localize")
    if "keywords" in jsonConfig:
        keywords = jsonConfig.get("keywords") or []
        Config.keywords = dict.fromkeys(keywords,True)
    if "maxCol" in jsonConfig:
        Config.maxCol = jsonConfig.get("maxCol")
    if "namespace" in jsonConfig:
        Config.namespace = jsonConfig.get("namespace")

def parse_csv_list(value):
    if value is None:
        return None
    items = [x.strip() for x in str(value).split(",") if x.strip()]
    return items

def parse_json_arg(value, optName):
    if value is None:
        return None
    try:
        return json.loads(value)
    except Exception:
        raise Exception("invalid json for %s: %s" % (optName, value))

def cfg_bool_choice(value):
    return "true" if value else "false"

def overlay_config(jsonConfig, options):
    """CLI 显式传入的项覆盖 --config。未传为 None，不覆盖。"""
    cfg = dict(jsonConfig or {})
    if options.inputDir is not None:
        cfg["input"] = options.inputDir
    if options.outputDir is not None:
        cfg["output"] = options.outputDir
    outputFormats = parse_csv_list(options.outputFormats)
    if outputFormats is not None:
        cfg["outputFormats"] = outputFormats
    keywords = parse_csv_list(options.keywords)
    if keywords is not None:
        cfg["keywords"] = keywords
    tags = parse_csv_list(options.exportTags)
    if tags is not None:
        cfg["tags"] = tags
    exclude = parse_csv_list(options.exclude)
    if exclude is not None:
        cfg["exclude"] = exclude
    if options.namespace is not None:
        cfg["namespace"] = options.namespace
    if options.indent is not None:
        cfg["indent"] = options.indent
    if options.constraintSeperator is not None:
        cfg["constraintSeperator"] = options.constraintSeperator
    if options.maxCol is not None:
        cfg["maxCol"] = options.maxCol
    if options.i18nDirectory is not None:
        cfg["i18nDirectory"] = options.i18nDirectory
    if options.i18nLanguage is not None:
        cfg["i18nLanguage"] = options.i18nLanguage
    if options.i18nExtension is not None:
        cfg["i18nExtension"] = options.i18nExtension
    if options.i18nSeperator is not None:
        cfg["i18nSeperator"] = options.i18nSeperator
    if options.genMeta is not None:
        cfg["genMeta"] = choice_to_bool(options.genMeta)
    if options.genMetaDetail is not None:
        cfg["genMetaDetail"] = choice_to_bool(options.genMetaDetail)
    if options.genMetaHeader is not None:
        cfg["genMetaHeader"] = choice_to_bool(options.genMetaHeader)
    if options.pretty is not None:
        cfg["pretty"] = choice_to_bool(options.pretty)
    if options.localize is not None:
        cfg["localize"] = choice_to_bool(options.localize)
    if options.classNameFirstUpper is not None:
        cfg["classNameFirstUpper"] = choice_to_bool(options.classNameFirstUpper)
    if options.fieldNameFirstUpper is not None:
        cfg["fieldNameFirstUpper"] = choice_to_bool(options.fieldNameFirstUpper)
    if options.i18nExportOneFile is not None:
        cfg["i18nExportOneFile"] = choice_to_bool(options.i18nExportOneFile)
    defaults = parse_json_arg(options.defaultsJson, "--defaults")
    if defaults is not None:
        cfg["defaults"] = defaults
    merge = parse_json_arg(options.mergeJson, "--merge")
    if merge is not None:
        cfg["merge"] = merge
    return cfg

def parser_for_format(outputFormat):
    if outputFormat == "lua":
        return Xls2LuaParser
    if outputFormat == "luacvs":
        return Xls2LuaCvsParser
    if outputFormat == "py":
        return Xls2PyParser
    if outputFormat == "json":
        return Xls2JsonParser
    if outputFormat == "binary":
        return Xls2BinaryParser
    if outputFormat == "csharp":
        return Xls2CSharpParser
    if outputFormat == "go":
        return Xls2GoParser
    if outputFormat == "schema":
        return Xls2SchemaParser
    return None

def main():
    usage = \
"""usage: python %prog [options]
e.g:
    python %prog --config=config.json
    python %prog --input=../Excel --output=../Output/Client --output-formats=csharp,binary,json,schema --tags=c
    python %prog --from-schema=schema.json --from-json=data.json --out=out.xlsx
    python %prog --from-schema=schema.json --from-binary=data.bytes --out=out.xlsx
    python %prog --from-dir=../Output/Client --export-xlsx=../Excel
    python %prog --import-xlsx=a.xlsx --schema-dir=schema --json-dir=json"""
    parser = optparse.OptionParser(usage=usage,version="%prog 0.0.1")
    parser.add_option("-c","--config",help="[optional] json config file; CLI args override file values")
    parser.add_option("-x","--onlyExportChange",action="store_true",default=False,help="[optional] only export change files")
    parser.add_option("--from-schema", dest="fromSchema", help="schema json for export")
    parser.add_option("--from-json", dest="fromJson", help="data json for export to excel")
    parser.add_option("--from-binary", dest="fromBinary", help="data binary for export to excel")
    parser.add_option("--out", dest="outPath", help="output xlsx path for export")
    parser.add_option("--from-dir", dest="fromDir", help="batch export root containing schema/ and json/")
    parser.add_option("--export-xlsx", dest="exportXlsx", help="batch export xlsx output dir")
    parser.add_option("--import-xlsx", dest="importXlsx", help="import workbook to schema+json dirs")
    parser.add_option("--schema-dir", dest="schemaDir", help="schema output dir for --import-xlsx")
    parser.add_option("--json-dir", dest="jsonDir", help="json output dir for --import-xlsx")
    parser.add_option("--sheet", dest="sheetName", help="optional sheet name for --import-xlsx")
    parser.add_option("--tags", dest="exportTags", help="tag filter, comma-separated (e.g. c,s); default empty = all columns")
    # optparse has no required=; input/output/outputFormats are checked after overlay.
    # overlay uses is not None, so do not set default= on these dests (would always override --config).
    parser.add_option("--input", dest="inputDir", help="excel input dir (config.input); required unless --config")
    parser.add_option("--output", dest="outputDir", help="generated config output dir (config.output); required unless --config")
    parser.add_option("--output-formats", dest="outputFormats", help="comma-separated formats: lua,luacvs,go,csharp,py,json,binary,schema; required unless --config")
    parser.add_option("--keywords", dest="keywords", help="comma-separated reserved field names; default matches Config.keywords")
    parser.add_option("--namespace", dest="namespace", help="code namespace, default %s" % Config.namespace)
    parser.add_option("--gen-meta", dest="genMeta", help="true/false or 1/0, generate field comments, default %s" % cfg_bool_choice(Config.genMeta), **BOOL_OPTION)
    parser.add_option("--gen-meta-detail", dest="genMetaDetail", help="true/false or 1/0, include excel remarks, default %s" % cfg_bool_choice(Config.genMetaDetail), **BOOL_OPTION)
    parser.add_option("--gen-meta-header", dest="genMetaHeader", help="true/false or 1/0, comments in file header, default %s" % cfg_bool_choice(Config.genMetaHeader), **BOOL_OPTION)
    parser.add_option("--pretty", dest="pretty", help="true/false or 1/0, pretty-print output, default %s" % cfg_bool_choice(Config.pretty), **BOOL_OPTION)
    parser.add_option("--localize", dest="localize", help="true/false or 1/0, translate i18nstring, default %s" % cfg_bool_choice(Config.localize), **BOOL_OPTION)
    parser.add_option("--class-name-first-upper", dest="classNameFirstUpper", help="true/false or 1/0, default %s" % cfg_bool_choice(Config.classNameFirstUpper), **BOOL_OPTION)
    parser.add_option("--field-name-first-upper", dest="fieldNameFirstUpper", help="true/false or 1/0, default %s" % cfg_bool_choice(Config.fieldNameFirstUpper), **BOOL_OPTION)
    parser.add_option("--i18n-directory", dest="i18nDirectory", help="i18n output directory, default %s" % Config.i18nDirectory)
    parser.add_option("--i18n-language", dest="i18nLanguage", help="i18n language, default %s" % Config.i18nLanguage)
    parser.add_option("--i18n-extension", dest="i18nExtension", help="i18n file ext: .po/.txt/.lua, default %s" % Config.i18nExtension)
    parser.add_option("--i18n-export-one-file", dest="i18nExportOneFile", help="true/false or 1/0, default false", **BOOL_OPTION)
    parser.add_option("--i18n-seperator", dest="i18nSeperator", help="txt i18n separator, default %s" % Config.i18nSeperator)
    parser.add_option("--exclude", dest="exclude", help="comma-separated excel files to skip")
    parser.add_option("--indent", dest="indent", help="indent string for pretty output, default 4 spaces")
    parser.add_option("--constraint-seperator", dest="constraintSeperator", help="constraint separator, default %s" % Config.constraintSeperator)
    parser.add_option("--max-col", dest="maxCol", type="int", help="max excel columns, default %s" % Config.maxCol)
    parser.add_option("--defaults", dest="defaultsJson", help='json object, e.g. {"int":0}')
    parser.add_option("--merge", dest="mergeJson", help='json object, e.g. {"toSheetName":["fromSheetName"]}')
    options,args = parser.parse_args()

    export_tags = None
    if options.exportTags:
        export_tags = [t.strip() for t in options.exportTags.split(",") if t.strip()]

    # --- import / export modes (no --config required) ---
    if options.fromDir:
        if not options.exportXlsx:
            parser.error("--export-xlsx required with --from-dir")
        export_dir(options.fromDir, options.exportXlsx, tags=export_tags)
        return

    if options.fromSchema and (options.fromJson or options.fromBinary):
        if not options.outPath:
            parser.error("--out required with --from-schema")
        if options.fromJson and options.fromBinary:
            parser.error("use either --from-json or --from-binary, not both")
        path = export_one(
            options.fromSchema,
            options.outPath,
            json_path=options.fromJson,
            binary_path=options.fromBinary,
            tags=export_tags,
        )
        print(json.dumps({"ok": True, "path": str(path)}, ensure_ascii=False))
        return

    if options.importXlsx:
        if not options.schemaDir or not options.jsonDir:
            parser.error("--schema-dir and --json-dir required with --import-xlsx")
        import_xlsx_to_dirs(
            options.importXlsx,
            options.schemaDir,
            options.jsonDir,
            options.sheetName,
        )
        return

    if options.fromSchema and not options.fromJson and not options.fromBinary:
        parser.error("--from-json or --from-binary required with --from-schema")
        return

    # --- forward gen: --config optional, CLI overrides file ---
    jsonConfig = {}
    if options.config is not None:
        configFileName = options.config
        if not configFileName.endswith(".json"):
            parser.error("config file need json")
            return
        fp = open(configFileName,"r",encoding="utf-8")
        txtConfig = fp.read()
        fp.close()
        jsonConfig = json.loads(txtConfig)
    try:
        jsonConfig = overlay_config(jsonConfig, options)
    except Exception as e:
        parser.error(str(e))
        return
    inputDir = jsonConfig.get("input")
    if not inputDir:
        parser.error("'input' required (--input or config.input)")
    outputDir = jsonConfig.get("output")
    if not outputDir:
        parser.error("'output' required (--output or config.output)")
    outputFormats = jsonConfig.get("outputFormats")
    if not outputFormats:
        parser.error("'outputFormats' required (--output-formats or config.outputFormats)")
    onlyExportChange = options.onlyExportChange
    exportFileList = []
    if onlyExportChange:
        result = subprocess.check_output(["svn", "status", inputDir], universal_newlines=True)
        lines = result.splitlines()
        for line in lines:
            lst = line.split(maxsplit=1)
            tag,fileName = lst[0],lst[1]
            if tag == "A" or tag == "M" or tag == "?":
                fileName = fileName.replace("\\","/")
                exportFileList.append(fileName)
    i18nExportOneFile = jsonConfig.get("i18nExportOneFile")
    i18nDirectory = jsonConfig.get("i18nDirectory")
    if i18nDirectory is None:
        i18nDirectory = Config.i18nDirectory
    i18nLanguage = jsonConfig.get("i18nLanguage")
    if i18nLanguage is None:
        i18nLanguage = Config.i18nLanguage
    i18nExtension = jsonConfig.get("i18nExtension") or Config.i18nExtension
    i18nSeperator = jsonConfig.get("i18nSeperator") or Config.i18nSeperator
    exclude = jsonConfig.get("exclude") or []

    apply_json_config(jsonConfig)

    # 载入本地化文本
    if i18nLanguage:
        readI18nFile(i18nExportOneFile,i18nDirectory,i18nLanguage,i18nExtension,i18nSeperator)
    sheets = getSheets()
    # Load types for list<Class> etc.; __class__.xlsx is not a data table (skipped below).
    # Schema format writes kind=0 class files separately via write_class_schemas_from_excel.
    readClass(inputDir)
    # 载入所有表
    for root,dirs,files in os.walk(inputDir):
        for dirname in dirs:
            for outputFormat in outputFormats:
                outputPath = os.path.join(outputDir,outputFormat,dirname)
                if not os.path.exists(outputPath):
                    os.makedirs(outputPath)
        for fileName in files:
            if fileName.startswith("~$"):
                # 临时文件
                continue
            if fileName.startswith("__class__"):
                # 类定义文件（类型已由 readClass 载入，不作为配表导出行）
                continue
            fullFileName = root + "/" + fileName
            if onlyExportChange and fullFileName not in exportFileList:
                continue
            relativeDirName = os.path.relpath(root,inputDir)
            if relativeDirName != "" and relativeDirName != ".":
                fullFileName = os.path.join(relativeDirName,fileName)
            else:
                fullFileName = fileName
            if fullFileName in exclude:
                continue
            wb = load_workbook(filename = os.path.join(root,fileName),data_only=True)
            for sheetIndex, sheetName in enumerate(wb.sheetnames):
                if not is_importable_sheet_title(sheetName):
                    continue
                sheet = Sheet(wb[sheetName],fullFileName,sheetName,sheetIndex)
                sheets[sheet.filename] = sheet
        # 表的引用检查(当onlyExportChange存在时,检查引用可能失效,此时你应该采用导出全表方式)
    if not onlyExportChange:
        for sheetName,sheet in sheets.items():
            sheet.checkRef()
    # 拆分表
    splitSheets = {}
    for sheetName,sheet in sheets.items():
        if sheet.splitCol != -1:
            splitSheets[sheet.filename] = sheet
    for sheetName,sheet in splitSheets.items():
        sheets.pop(sheetName)
        tempSheets = sheet.splitSheets()
        for tempSheet in tempSheets:
            sheets[tempSheet.filename] = tempSheet
            print("splitSheet,filename=%s,maxCol=%s,dataRow=%s,idCol=%d" % (tempSheet.filename,tempSheet.maxCol,tempSheet.dataRow,tempSheet.idCol))

    # 合并表
    merge = jsonConfig.get("merge")
    if merge != None:
        for toFileName,fromFileNames in merge.items():
            toSheet = sheets.get(toFileName)
            if toSheet != None:
                for fromFileName in fromFileNames:
                    fromSheet = sheets.get(fromFileName)
                    if fromSheet != None:
                        sheets.pop(fromFileName)
                        toSheet.mergeFrom(fromSheet)

    XlsParser.ignoreGenTables = onlyExportChange
    for outputFormat in outputFormats:
        print("xls2cfg,outputDir=%s,outputFormat=%s" % (outputDir,outputFormat))
        Parser = parser_for_format(outputFormat)
        if Parser is None:
            raise Exception("unknown outputFormat: %s" % outputFormat)

        # 生成配置文件
        output = os.path.join(outputDir,outputFormat)
        for sheetName,sheet in sheets.items():
            parser = Parser(sheet,output)
            parser.parse()
        if outputFormat == "schema":
            write_class_schemas_from_excel(inputDir, output)
        Parser.endParse(output)

    # 生成国际化待翻译文本(如果目标文件已存在,则合并翻译文本)
    if not onlyExportChange and i18nLanguage:
        writeI18nFile(i18nExportOneFile,i18nDirectory,i18nLanguage,i18nExtension,i18nSeperator)

if __name__ == "__main__":
    main()
