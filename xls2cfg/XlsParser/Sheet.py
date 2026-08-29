#coding:utf-8
#@author sundream
#@date 2025-09-10

import os
import time
from openpyxl.utils import get_column_letter
from XlsParser.CheckType import toValue, intLiteralBase, INT_BASE_TYPENAMES
from XlsParser.Type import Type, parse_tags_cell
from XlsParser import Convert
from XlsParser.Config import Config
from copy import deepcopy

sheets = {}

def getSheets():
    return sheets


def split_at_display(title):
    """Split 'name@displayName' → (name, displayName|None).

    Used for workbook stems (effect@效果) and sheet titles (Buff@增益).
    """
    if title is None:
        return None, None
    s = str(title)
    if "@" in s:
        name, disp = s.split("@", 1)
        name = name.strip()
        disp = disp.strip()
        return name, disp if disp else None
    return s, None


def is_importable_sheet_title(sheet_title):
    """Skip internal sheets; allow SheetName@中文备注 (ascii check on name only)."""
    logical, _ = split_at_display(sheet_title)
    if not logical:
        return False
    if logical.startswith("_") or logical.startswith("Sheet"):
        return False
    if not logical.isascii():
        return False
    return True


def compose_table_display_name(sheet):
    """Table schema displayName.

    - Sheet title ``data@英雄`` / ``Buff@增益`` → ``英雄`` / ``增益`` (only the part after @)
    - Sheet title ``data`` (no @) + workbook ``hero@英雄`` → ``英雄`` (workbook @ part)
    - Other sheets without @ → ``{workbook显示名}-{sheetName}`` when workbook has display, else sheetName
    """
    if sheet is None:
        return None
    wb = (getattr(sheet, "workbookDisplayName", None) or "").strip()
    sheet_disp = (getattr(sheet, "sheetDisplayName", None) or "").strip()
    sheet_name = (getattr(sheet, "sheetName", None) or "").strip()
    # Sheet @display wins alone (do not prefix workbook display).
    if sheet_disp:
        return sheet_disp
    if not sheet_name:
        return wb or None
    # Classic 1-file-1-table: exact sheet title "data" (case-sensitive).
    if sheet_name == "data" and wb:
        return wb
    if wb:
        return "%s-%s" % (wb, sheet_name)
    return sheet_name


class MergeCell(object):
    # list<any> 和 map<string,any>类型允许拆分单元格填值
    def __init__(self,startCol,count):
        self.startCol = startCol                # 起始列号(0-based)
        self.count = count                      # 占据单元格数
        self.type = None                        # 类型
        self.mapkeys = {}                       # key字典: key -> True
        self.fieldCount = None                  # 元素字段数

    def setType(self,type):
        self.type = type
        if self.type.typename == "list" and self.type.valueType.isClass():
            self.fieldCount = len(self.type.valueType.fields)

class Sheet(object):
    def __init__(self,sheet,xlsFilename,sheetName,sheetIndex=0):
        self.xlsFilename = os.path.basename(xlsFilename)
        self.sheetIndex = sheetIndex
        # sheet title may be SheetName@DisplayName
        self.sheetName, self.sheetDisplayName = split_at_display(sheetName)
        fileStem, workbookDisplay = split_at_display(os.path.splitext(self.xlsFilename)[0])
        self.workbookName = fileStem
        self.workbookDisplayName = workbookDisplay
        # legacy: comment = workbook display (file @ part)
        self.comment = workbookDisplay
        # logical table name / output stem
        if self.sheetName == "data":
            self.filename = self.workbookName
        else:
            self.filename = self.workbookName + "_" + self.sheetName
        self.defaults = deepcopy(Config.defaults)
        for k,v in Type.basicTypes.items():
            if not self.defaults.get(k):
                self.defaults[k] = v
        for k,v in Type.containerTypes.items():
            if not self.defaults.get(k):
                self.defaults[k] = v
        self.constraintSeperator = Config.constraintSeperator
        self.sheet = sheet
        # 前5行为头部信息(1-based),1=描述,2=字段name,3=类型,4=约束,5=tag标签
        self.headerRow = 5              # 表头5行
        self.col2desc = {}              # 列号 -> 描述(0-based)
        self.col2comment = {}           # 列号 -> 批注
        self.col2key = {}               # 列号 -> 变量名
        self.col2type = {}              # 列号 -> 类型
        self.col2constraint = {}        # 列号 -> 约束
        self.col2tags = {}              # 列号 -> [tag,...]（不含 group=）
        self.col2group = {}             # 列号 -> 编辑器分组名
        self.idCol = 0                  # id列
        self.splitCol = -1              # 拆分列
        self.splitIdCol = -1            # 拆分主键列
        self.col2uniques = {}           # 列唯一性检查表
        self.key2col = {}               # 变量名 -> 列号,方便引用检查
        self.rows = []                  # 表格数据
        self.maxRow = 0                 # 最大行号
        self.maxCol = 0                 # 最大列号
        self.dataRow = 0                # 有效数据行数(不包括表头行)
        self.mergeCells = {}            # 列号 -> 合并单元格信息{col: 合并到单元格的列号, type: 类型, count=合并单元格数, mapkeys=固定key字典}
        self.col2mapkey = {}            # 列号 -> mapkey
        self.col2intBase = {}           # 列号 -> 16/2（整列统一进制，字段级）
        if self.sheet == None:
            return
        self.maxRow = self.sheet.max_row
        self.maxCol = self.sheet.max_column
        self.singleton = False          # True=单例表
        for firstRow in self.sheet.iter_rows(min_row=1,max_row=self.headerRow):
            for j in range(0,self.maxCol):
                cell = firstRow[j]
                if cell.value:
                    if cell.value.startswith("#"):
                        if cell.value == "##end":
                            self.maxCol = j
                        elif cell.value == "##key" and j == 0:
                            self.singleton = True
                        else:
                            self.col2tags[j] = ["__ignore"]
            break
        self._loadStartTime = time.perf_counter()
        print("loadSheetBegin,xlsFilename=%s,sheetName=%s" % (self.xlsFilename,self.sheetName))
        if self.maxCol > Config.maxCol:
            raise Exception("maxCol=%d > Config.maxCol=%d" % (self.maxCol,Config.maxCol))
        if self.singleton:
            self.initSingletonSheet()
            return
        # 依据第3行合并的单元格信息确定哪些数据需要合并单元格
        for mergedCell in self.sheet.merged_cells.ranges:
            # 1-based
            startCol,startRow,endCol,endRow = mergedCell.bounds
            if startCol > self.maxCol:
                continue
            assert(startRow == endRow)
            if startRow > 3:
                raise("mergeCell only allow in first 3 rows")
            startCol -= 1
            if startRow == 3:
                self.mergeCells[startCol] = MergeCell(startCol,endCol-startCol)
                for i in range(startCol,endCol):
                    self.mergeCells[i] = self.mergeCells[startCol]

        i = 0
        for row in self.sheet.iter_rows(min_row=1,max_row=self.headerRow):
            i = i + 1
            for j in range(0,self.maxCol):
                cell = row[j]
                if i == 1:
                    # Keep newlines for schema round-trip; genMeta flattens later.
                    if cell.value is None:
                        self.col2desc[j] = None
                    else:
                        self.col2desc[j] = cell.value
                    if cell.comment is None:
                        self.col2comment[j] = None
                    else:
                        self.col2comment[j] = cell.comment.text
                elif i == 2:
                    name = cell.value
                    if name is not None:
                        name = name.strip()
                        if Config.isKeyword(name):
                            raise Exception(self._message(i,j,name + " is keywords"))
                        if name in self.key2col:
                            raise Exception(self._message(i,j,"repeat key=%s" % name))
                        self.col2key[j] = name
                        self.key2col[name] = j
                    else:
                        self.col2key[j] = None
                elif i == 3:
                    self.col2type[j] = None
                    typename = cell.value
                    if typename is None:
                        if self.col2key.get(j) is None:
                            # 空列,可能是备注列
                            continue
                        raise Exception(self.message(i,j,"typename required"))
                    try:
                        typ = Type.getOrCreate(typename)
                    except Exception as e:
                        raise Exception(self.message(i,j,e))
                    self.col2type[j] = typ
                    if j in self.mergeCells:
                        self.mergeCells[j].setType(typ)
                elif i == 4:
                    self.col2constraint[j] = {}
                    typ = self.col2type[j]
                    raw = cell.value
                    if raw is not None and str(raw).strip() != "":
                        kvs = str(raw).split(self.constraintSeperator)
                        for kv in kvs:
                            kv = kv.strip()
                            if not kv:
                                continue
                            lst = kv.split("=")
                            k = lst[0].strip()
                            v = None
                            if len(lst) > 1:
                                v = lst[1]
                            if not k:
                                continue
                            if k == "convert" or k == "ref":
                                if v is None:
                                    raise Exception(self._message(i,j,"expire format '%s=value'" % k))
                                self.col2constraint[j][k] = v
                            elif k == "min" or k == "max":
                                if v is None:
                                    raise Exception(self._message(i,j,"expire format '%s=value'" % k))
                                self.col2constraint[j][k] = int(v)
                            elif k == "default":
                                if v is None:
                                    raise Exception(self._message(i,j,"expire format '%s=value'" % k))
                                # allow default=nil or default=None
                                if v == "nil" or v == "None":
                                    v = None
                                else:
                                    ok,v = toValue(v,typ)
                                    if not ok:
                                        errMsg = v
                                        raise Exception(self._message(i,j,errMsg))
                                self.col2constraint[j][k] = v
                            elif k == "unique":
                                self.col2constraint[j][k] = True
                            elif k == "not_null":
                                self.col2constraint[j][k] = True
                            elif k == "not_localize":
                                self.col2constraint[j][k] = True
                            elif k == "split":
                                try:
                                    # 1-based
                                    v = int(v) - 1
                                except Exception as e:
                                    if v not in self.key2col:
                                        raise Exception("invalid split column: %s" % v)
                                    v = self.key2col[v]
                                self.col2constraint[j][k] = v
                                if self.splitCol != -1 and self.splitCol != j:
                                    raise Exception("multiple split constraint")
                                self.splitCol = j
                                self.splitIdCol = v
                                if self.splitCol == self.splitIdCol:
                                    raise Exception("split column == id column")
                                if self.splitIdCol < 0:
                                    raise Exception("split id column < 0")
                            elif k == "limit":
                                if v is None:
                                    raise Exception(self._message(i,j,"expire format '%s=value'" % k))
                                self.col2constraint[j][k] = eval(v)
                            elif k.startswith("."):
                                mapkey = k[1:]
                                self.col2mapkey[j] = mapkey
                                mergeCell = self.mergeCells[j]
                                mergeCell.mapkeys[mapkey] = True
                                mergeColType = mergeCell.type
                                if mergeColType.typename != "map":
                                    mergeColType = mergeColType.valueType
                                if mergeColType.typename != "map" and not mergeColType.isClass():
                                    raise Exception(self._message(i,j,"invalid mapkey '%s',merge cell type exepect map/class,but got '%s'" % (k,mergeColType)))
                            else:
                                raise Exception(self._message(i,j,"invalid constraint,value='%s',constraint seperator is '%s'" % (cell.value,self.constraintSeperator)))
                    if self.getConstraint(j,"unique") and not self.getConstraint(j,"not_null"):
                        # unique列自带not_null约束（仅校验；导出 schema 不写，见 _not_null_implicit）
                        self.col2constraint[j]["not_null"] = True
                        self.col2constraint[j]["_not_null_implicit"] = True
                elif i == 5:
                    mergeCell = self.mergeCells.get(j)
                    if mergeCell and len(mergeCell.mapkeys) > 0:
                        mergeCell.fieldCount = len(mergeCell.mapkeys)
                    if cell.value:
                        try:
                            tags, group = parse_tags_cell(cell.value)
                        except Exception as e:
                            raise Exception(self.message(i,j,e))
                        if group:
                            self.col2group[j] = group
                        if tags:
                            if j not in self.col2tags or self.col2tags[j] is None:
                                self.col2tags[j] = []
                            self.col2tags[j].extend(tags)
                        elif j not in self.col2tags:
                            self.col2tags[j] = None
                    elif j not in self.col2tags:
                        self.col2tags[j] = None

        # id 列默认 unique（仅校验）；Excel 未写 unique 时不导出到 schema
        if self.splitCol == -1 and not self.getConstraint(0,"unique"):
            self.col2constraint[0]["unique"] = True
            self.col2constraint[0]["_unique_implicit"] = True
            if not self.getConstraint(0,"not_null"):
                self.col2constraint[0]["not_null"] = True
                self.col2constraint[0]["_not_null_implicit"] = True

        for i,row in enumerate(self.sheet.iter_rows(min_row=self.headerRow+1,max_row=self.maxRow,values_only=True)):
            if row[0] is None or (type(row[0]) == str and row[0].isspace()):
                # 遇到空行则结束
                break
            if type(row[0]) == str and row[0].startswith("#"):
                # 忽略注释行
                continue
            self.dataRow = self.dataRow + 1
            line = []
            self.rows.append(line)
            # intBase：仅看首个数据行（与 toInt 相同的 0x/0b 规则；不含 bigint）
            if self.dataRow == 1:
                for j in range(0, self.maxCol):
                    typ = self.getColType(j)
                    if typ is None or typ.typename not in INT_BASE_TYPENAMES:
                        continue
                    base = intLiteralBase(row[j])
                    if base:
                        self.col2intBase[j] = base
            # 约束检查
            for j in range(0,self.maxCol):
                value = row[j]
                not_null = self.getConstraint(j,"not_null")
                if not_null and value is None:
                    raise Exception(self.message(i,j,"can not be null"))
                typ = self.getColType(j)
                if typ is None:
                    # 空列/备注列
                    line.append(value)
                    continue
                localize = Config.localize
                if self.getConstraint(j,"not_localize"):
                    localize = False
                ok,errMsg = toValue(value,typ,{
                    "xlsFilename":self.filename,
                    "row":i,
                    "col":j,
                    "headerRow":self.headerRow,
                    "localize":localize,
                    "depth":0,
                })
                if not ok:
                    raise Exception(self.message(i,j,errMsg))
                else:
                    value = errMsg
                typename = typ.typename
                if value is None:
                    value = self.getDefault(i,j)
                else:
                    convert = self.getConvert(j)
                    if convert:
                        value = convert(value)
                unique = self.getConstraint(j,"unique")
                if unique:
                    if j not in self.col2uniques:
                        uniques = {}
                        self.col2uniques[j] = uniques
                    else:
                        uniques = self.col2uniques[j]
                    if value in uniques:
                        raise Exception(self.message(i,j,"not unique,conflict with cell %s" % (self.cellname(uniques[value],j))))
                    uniques[value] = self.headerRow + i
                minValue = self.getConstraint(j,"min")
                if minValue is not None and value < minValue:
                    raise Exception(self.message(i,j,"min=%s,value=%s" % (minValue,value)))
                maxValue = self.getConstraint(j,"max")
                if maxValue is not None and value > maxValue:
                    raise Exception(self.message(i,j,"max=%s,value=%s" % (maxValue,value)))
                limit = self.getConstraint(j,"limit")
                if limit is not None and value not in limit:
                    raise Exception(self.message(i,j,"limit=%s,value=%s" % (limit,value)))
                if j in self.mergeCells:
                    mergeCell = self.mergeCells[j]
                    mergeCellType = mergeCell.type
                    mergeStartCol = mergeCell.startCol
                    if mergeCellType.typename == "list":
                        if j == mergeStartCol:
                            lst = []
                            line.append(lst)
                        else:
                            lst = line[mergeStartCol]
                        if j in self.col2mapkey:
                            # type: list<map<string,any>>  or list<Class>
                            mapkey = self.col2mapkey[j]
                            assert(mergeCellType.valueType.typename == "map" or mergeCellType.valueType.isClass())
                            fieldCount = mergeCell.fieldCount
                            elemIdx = (j - mergeStartCol) // fieldCount
                            if (j - mergeStartCol) % fieldCount == 0:
                                map = {}
                                lst.append(map)
                            else:
                                map = lst[elemIdx]
                            map[mapkey] = value
                        else:
                            # type: list<any>
                            lst.append(value)
                    elif mergeCellType.typename == "map":
                        # type: map<string,any>
                        if j == mergeStartCol:
                            map = {}
                            line.append(map)
                        else:
                            map = line[mergeStartCol]
                        mapkey = self.col2mapkey[j]
                        map[mapkey] = value
                    if j != mergeStartCol:
                        # 空值占位
                        line.append(None)
                else:
                    line.append(value)

        elapsedMs = (time.perf_counter() - self._loadStartTime) * 1000
        print("loadSheetEnd,xlsFilename=%s,sheetName=%s,maxRow=%d,maxCol=%d,dataRow=%d,elapsedMs=%.2fms" % (self.xlsFilename,self.sheetName,self.maxRow,self.maxCol,self.dataRow,elapsedMs))

    def initSingletonSheet(self):
        # 首行为表头行,固定名字为: key | type | value | tags | desc
        self._loadStartTime = time.perf_counter()
        print("loadSingleSheetBegin,xlsFilename=%s,sheetName=%s" % (self.xlsFilename,self.sheetName))
        self.headerRow = 1
        self.header = {}            # col -> keyword
        for row in self.sheet.iter_rows(min_row=1,max_row=self.headerRow):
            for j in range(0,self.maxCol):
                cell = row[j]
                value = cell.value
                if value is None:
                    continue
                if value == "##key":
                    self.header[j] = "key"
                elif value == "type":
                    self.header[j] = "type"
                elif value == "value":
                    self.header[j] = "value"
                elif value == "tags":
                    self.header[j] = "tags"
                elif value == "desc":
                    self.header[j] = "desc"
        line = []
        self.rows.append(line)
        for i,row in enumerate(self.sheet.iter_rows(min_row=self.headerRow+1,max_row=self.maxRow,values_only=True)):
            if row[0] is None or (type(row[0]) == str and row[0].isspace()):
                # 遇到空行则结束
                break
            if type(row[0]) == str and row[0].startswith("#"):
                # 忽略注释行
                self.col2type[i] = None
                line.append(None)
                continue
            for j in range(0,self.maxCol):
                value = row[j]
                self.col2comment[i] = None
                if self.header[j] == "key":
                    if Config.isKeyword(value):
                        raise Exception(self.message(i,j,value + " is keywords"))
                    if value in self.key2col:
                        raise Exception(self.message(i,j,"repeat key=%s" % value))
                    self.col2key[i] = value
                    self.key2col[value] = j
                elif self.header[j] == "type":
                    try:
                        typ = Type.getOrCreate(value)
                    except Exception as e:
                        raise Exception(self.message(i,j,e))
                    self.col2type[i] = typ
                elif self.header[j] == "value":
                    if value is None:
                        value = self.getDefault(j,i)
                    if typ is not None and typ.typename in INT_BASE_TYPENAMES:
                        base = intLiteralBase(value)
                        if base:
                            self.col2intBase[i] = base
                    ok,value = toValue(value,typ)
                    if not ok:
                        errMsg = value
                        raise Exception(self.message(i,j,errMsg))
                    line.append(value)
                elif self.header[j] == "tags":
                    if value:
                        try:
                            tags, group = parse_tags_cell(value)
                        except Exception as e:
                            raise Exception(self.message(i,j,e))
                        self.col2tags[i] = tags
                        if group:
                            self.col2group[i] = group
                    else:
                        self.col2tags[i] = None
                elif self.header[j] == "desc":
                    self.col2desc[i] = value
        self.dataRow = 1
        self.maxCol = self.maxRow
        self.maxRow = self.headerRow + 1    # 1 row
        elapsedMs = (time.perf_counter() - self._loadStartTime) * 1000
        print("loadSingleSheetEnd,xlsFilename=%s,sheetName=%s,maxRow=%d,maxCol=%d,dataRow=%d,elapsedMs=%.2fms" % (self.xlsFilename,self.sheetName,self.maxRow,self.maxCol,self.dataRow,elapsedMs))


    def message(self,row,col,msg):
        return self._message(self.headerRow+row,col,msg)

    def _message(self,row,col,msg):
        return "xlsFilename=%s,sheetName=%s,cell=%s,value=%s,msg=%s" % (self.xlsFilename,self.sheetName,self.cellname(row,col),self.sheet.cell(row+1,col+1).value,msg)

    def cellname(self,row,col):
        return "%s%s" % (get_column_letter(col+1),row+1)

    def value(self,row,col):
        val = self.rows[row][col]
        return val

    # def row(self,i):
    #     return self.rows[i]

    def getColType(self,col):
        if col in self.mergeCells:
            mergeCell = self.mergeCells[col]
            typ = mergeCell.type
            valueType = typ.valueType
            if valueType.typename == "map":
                # list<map<keytype,valuetype>>
                valueType = valueType.valueType
            elif valueType.isClass():
                # list<Class>
                if col in self.col2mapkey:
                    mapkey = self.col2mapkey[col]
                    for field in valueType.fields:
                        if field.name == mapkey:
                            valueType = field.type
            typ = valueType
        else:
            typ = self.col2type[col]
        return typ

    def getDefault(self,row,col):
        typ = self.getColType(col)
        if typ.isEnum():
            typ = typ.underlyingType()
        typename = typ.typename
        constraint = self.col2constraint.get(col)
        if constraint and "default" in constraint:
            default = constraint["default"]
        else:
            default = self.defaults.get(typename)
        return default

    def getConvert(self,col):
        convert = self.getConstraint(col,"convert")
        if convert:
            return getattr(Convert,convert)

    def getConstraint(self,col,key):
        constraint = self.col2constraint.get(col)
        if not constraint:
            return None
        return constraint.get(key)

    @staticmethod
    def _ref_supported_leaf(typ):
        """True if typ can be a ref leaf (scalar / enum)."""
        if typ is None:
            return False
        if typ.typename in ("list", "map", "json"):
            return False
        if typ.isClass():
            return False
        return True

    def _ref_type_error(self, typ):
        """Return error message if typ cannot carry a column-level ref, else None."""
        if typ is None:
            return "ref column has no type"
        if typ.typename == "list":
            elem = typ.valueType
            if not self._ref_supported_leaf(elem):
                return "ref only supports list<scalar|enum>, got list<%s>" % (
                    elem.fullTypename if elem else "?")
            return None
        if typ.typename == "map":
            val_t = typ.valueType
            if not self._ref_supported_leaf(val_t):
                return "ref only supports map<*,scalar|enum>, got map<%s,%s>" % (
                    typ.keyType.fullTypename if typ.keyType else "?",
                    val_t.fullTypename if val_t else "?")
            return None
        if not self._ref_supported_leaf(typ):
            return "ref not supported on type '%s' (only scalar/enum, list<scalar|enum>, map<*,scalar|enum>)" % typ.fullTypename
        return None

    def _ref_values(self, value, typ):
        """Collect leaf values that must exist in the referenced column."""
        err = self._ref_type_error(typ)
        if err:
            return None, err
        if typ.typename == "list":
            return list(value or []), None
        if typ.typename == "map":
            return list((value or {}).values()), None
        return [value], None

    @staticmethod
    def _parse_ref(ref):
        """Parse ref target: 'table.field' or 'table' (field defaults to id).

        Returns (refFilename, refColname) or (None, errorMessage).
        """
        if not ref or not str(ref).strip():
            return None, "reference constraint empty; expect 'ref=table' or 'ref=table.field'"
        ref = str(ref).strip()
        if "." in ref:
            parts = ref.split(".")
            if len(parts) != 2 or not parts[0].strip() or not parts[1].strip():
                return None, "reference constraint expire format 'ref=table' or 'ref=table.field',but got value='%s'" % ref
            return parts[0].strip(), parts[1].strip()
        return ref, "id"

    def checkRef(self):
        if self.singleton:
            return
        # Validate columns that declare ref (even if all cells empty).
        for j, constraint in self.col2constraint.items():
            if not constraint or not constraint.get("ref"):
                continue
            ref = constraint["ref"]
            typ = self.col2type.get(j)
            err = self._ref_type_error(typ)
            if err:
                raise Exception(self.message(0, j, err))
            refFilename, refColname = self._parse_ref(ref)
            if refFilename is None:
                raise Exception(self.message(0, j, refColname))
            refSheet = self.getSheet(refFilename)
            if refSheet is None:
                raise Exception(self.message(0, j, "reference's excel not exist,refFilename=%s,refColname=%s" % (refFilename, refColname)))
            if refSheet.key2col.get(refColname) is None:
                raise Exception(self.message(0, j, "reference's column not exist,refFilename=%s,refColname=%s" % (refFilename, refColname)))

        for i, row in enumerate(self.rows):
            for j, value in enumerate(row):
                if not value:
                    continue
                ref = self.getConstraint(j, "ref")
                if not ref:
                    continue
                refFilename, refColname = self._parse_ref(ref)
                if refFilename is None:
                    raise Exception(self.message(i, j, refColname))
                refSheet = self.getSheet(refFilename)
                refSheetCol = refSheet.key2col.get(refColname)
                values, err = self._ref_values(value, self.col2type[j])
                if err:
                    raise Exception(self.message(i, j, err))
                for item in values:
                    if item is None or item == "":
                        continue
                    ok = False
                    for refRow in refSheet.rows:
                        if item == refRow[refSheetCol]:
                            ok = True
                            break
                    if not ok:
                        raise Exception(self.message(i, j, "reference's value not exist,value=%s,refFilename=%s,refColname=%s" % (item, refFilename, refColname)))

    def mergeFrom(self,fromSheet):
        for row in fromSheet.rows:
            self.dataRow = self.dataRow + 1
            self.rows.append(row)

    def splitSheets(self):
        if self.splitCol == -1:
            return None
        sheetDict = {}
        for i in range(0,self.dataRow):
            row = self.rows[i]
            val = row[self.splitCol]
            sheet = sheetDict.get(val)
            if sheet is None:
                sheet = Sheet(None,self.xlsFilename,self.sheetName,self.sheetIndex)
                sheet.sheetDisplayName = self.sheetDisplayName
                sheet.workbookName = self.workbookName
                sheet.workbookDisplayName = self.workbookDisplayName
                sheet.comment = self.comment
                sheet.idCol = self.splitIdCol
                # 复制属性
                sheet.col2desc = self.col2desc
                sheet.col2comment = self.col2comment
                sheet.col2key = self.col2key
                sheet.col2type = self.col2type
                sheet.col2constraint = self.col2constraint
                sheet.col2tags = self.col2tags
                sheet.col2group = self.col2group
                sheet.col2uniques = self.col2uniques
                sheet.key2col = self.key2col
                sheet.splitCol = self.splitCol
                sheet.maxRow = self.maxRow
                sheet.maxCol = self.maxCol
                sheet.mergeCells = self.mergeCells
                sheet.col2mapkey = self.col2mapkey
                sheet.col2intBase = self.col2intBase

                sheet.filename += "_" + str(val)
                sheetDict[val] = sheet
            sheet.dataRow = sheet.dataRow + 1
            sheet.rows.append(row)
        return sheetDict.values()

    def getSheet(self,filename):
        global sheets
        return sheets.get(filename)